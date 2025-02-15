import math
import datetime
import openpyxl
import openpyxl.cell
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from src.utils.constants import *

class TxCheckManager:
    def __init__(self):
        self.workbook = openpyxl.load_workbook(filename = './templates/TxCheck.xlsx') # Apertura de la plantilla del FOR_Registro Monitoreo in Situ TDT.
        self.sheet = self.workbook[self.workbook.sheetnames[0]] # Selección de la hoja 'Registro Monitoreo in Situ TDT'.

    @staticmethod
    def _fill_color(cell: openpyxl.cell.cell.Cell, performance: float):
        """
        Aplica el color a una celda en función del desempeño.
        """
        if performance < 50:
            red, green = 'FF', f'{hex(int(5.1 * performance))[2:].zfill(2)}'
        elif performance > 50:
            green, red = 'FF', f'{hex(int(-5.1 * (performance - 100)))[2:].zfill(2)}'
        else:
            red, green = 'FF', 'FF'
        
        color = f'{red}{green}00'
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    @staticmethod
    def get_index_warnings(letter, index):
        """ Actualiza el índice y la letra de advertencias cuando el índice alcanza 8. """
        next_letters = {'A': 'D', 'D': 'G'}
        return (next_letters.get(letter, letter), 0) if index == 8 else (letter, index + 1)
    
    @staticmethod
    def calculate_points(value, key, worst, ideal, weight):
        """ Calcula los puntos basados en la diferencia con el valor ideal. """
        if isinstance(value, str):  # Si el valor es un string, usa el peor caso
            return 0, worst, True

        divisor = worst if key in ['LEVel', 'CFOFfset', 'BROFfset', 'IMBalance', 'QERRor'] else abs(ideal) - abs(worst)
        points = round(weight * (1 - abs(ideal - value) / divisor))
        
        return max(0, min(points, weight)), value, False  # Asegurar que los puntos estén en el rango permitido
    
    @staticmethod
    def calculate_performance(value, ideal, worst):
        """ Calcula el porcentaje de rendimiento. """
        performance = round(100 * (1 - (abs(ideal) - abs(value)) / (abs(ideal) - abs(worst))))
        return max(0, min(performance, 100))
    
    def fill_table(self, txcheck_dictionary: dict):
        index_table = 0
        letter_warnings = 'A'
        index_warnings = 0
        total_points = 0

        for key, (worst, ideal, weight, *extra) in TXCHECK_PARAMETERS.items():
            value = txcheck_dictionary[key]

            points, value, problematic = self.calculate_points(value, key, worst, ideal, weight)

            if problematic:
                print(key, value, 'Entré en el ciclo problemático')

            performance = self.calculate_performance(value, ideal, worst)

            if points == 0:
                self.sheet[f'C{49 + 2 * index_table}'] = 'Out of limit(s)'
                letter_warnings, index_warnings = self.get_index_warnings(letter_warnings, index_warnings)
                self.sheet[f'{letter_warnings}{224 + index_warnings}'] = extra[0] if extra else "Warning"

            # Llenado de la tabla
            self.sheet[f'L{48 + 2 * index_table}'] = value
            self.sheet[f'M{49 + 2 * index_table}'] = performance
            self.sheet[f'I{48 + 2 * index_table}'] = points

            self._fill_color(self.sheet[f'C{49 + 2 * index_table}'], performance)

            index_table += 1
            total_points += points

            total_performance = int(100*total_points/835)
            self._fill_color(self.sheet['D236'],total_performance)

    def fill_images(self, folder_path: str, channel: int):
        """ Carga, redimensiona e inserta imágenes en la plantilla según el canal. """
    
        # Lista de sufijos de archivos y sus posiciones en la plantilla
        image_positions = [
            ('003', 'A96'),  ('010', 'F96'),  ('001', 'A115'), ('012', 'F115'),
            ('013', 'A141'), ('014', 'F141'), ('008', 'A159'), ('009', 'F159'),
            ('004', 'A188'), ('005', 'F188'), ('006', 'A206'), ('011', 'F206')
        ]
        
        # Carga y redimensionamiento de imágenes
        images = {
            position: Image(f"{folder_path}/{TV_TABLE[channel]}_{suffix}.png")
            for suffix, position in image_positions
        }
        
        for img in images.values():
            img.width *= 0.27
            img.height *= 0.27

        # Inserción en la plantilla
        for position, img in images.items():
            self.sheet.add_image(img, position)

    def get_txchech_report(self, idn_string: str, txcheck_dictionary: dict, folder_path: str, channel: int):
        """ Retorna el reporte de TxCheck. """
        self.sheet[f'H9'] = idn_string.split(sep=',')[2]
        now = datetime.datetime.now() # Obetención de fecha y hora de la medida
        self.sheet[f'C3'] = f'{now.day}.{now.month}.{now.year}, {now.hour}:{now.minute}:{now.second}'
        self.sheet[f'D21'] = TV_TABLE[channel]
        self.sheet[f'D23'] = '75Ω'
        
        self.fill_table(txcheck_dictionary)
        self.fill_images(folder_path, channel)

        self.workbook.save(f'{folder_path}/txCheck1.xlsx')

if __name__ == '__main__':
    dic = {'SALower': 14.789680481, 'SAUPper': 8.43473815918, 'LEVel': 60.6906126339,
           'CFOFfset': -49.8, 'BROFfset': -0.11, 'PERatio': 0.0, 'BERLdpc': 0.0013,
           'BBCH': 0.0, 'FERatio': 0.0, 'ESRatio': 0.0, 'IMBalance': -0.3, 'QERRor': 0.01,
           'CSUPpression': 5.1, 'MRLO': 27.535, 'MPLO': -4.877, 'MRPLp': 28.257,
           'MPPLp': 2.1124245358, 'ERPLp': 2.53019894896, 'EPPLp': 51.3322602672,
           'AMPLitude': 48.2967691422, 'PHASe': 692.743804932, 'GDELay': 4.279262e-05}
    
    txcheck = TxCheckManager()
    txcheck.get_txchech_report('0,0,105406/013', dic, './tests', 16)