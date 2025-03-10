import openpyxl
from openpyxl.drawing.image import Image
import pandas as pd
from rapidfuzz import process, fuzz
import rasterio
from rasterio.plot import show
import numpy as np
import matplotlib.pyplot as plt
from shapely.geometry import LineString
from pyproj import Transformer
import geopandas as gpd
from shapely.geometry import Point, LineString
import datetime
from src.utils.constants import *
import os
import shutil

class ExcelReport:
    def __init__(self):
        self.analog_filename  = './templates/FOR_Registro Monitoreo In Situ TV Analógica_V0.xlsm'
        self.digital_filename = './templates/FOR_Registro Monitoreo In Situ TDT_V0.xlsm'

        # Creación de workbooks de ambos formatos
        self.wb_analog  = openpyxl.load_workbook(self.analog_filename)
        self.register_sheet = self.wb_analog[self.wb_analog.sheetnames[0]]
        self.graphical_supports_sheet = self.wb_analog[self.wb_analog.sheetnames[2]]

        self.wb_digital = openpyxl.load_workbook(self.digital_filename)
        self.general_info_sheet = self.wb_digital[self.wb_digital.sheetnames[1]]
        self.channel_template_sheet = self.wb_digital[self.wb_digital.sheetnames[2]]

        # Cargue de archivo de referencias
        self.filename_references = './src/utils/Referencias.xlsx'
        self.stations  = pd.read_excel(self.filename_references, sheet_name = 2)


    @staticmethod
    def _resize_image(img: Image):
        # Constante de conversión de píxeles a cm
        PIXEL_TO_CM = 0.02646

        # Obtener dimensiones originales en píxeles
        original_width = img.width
        original_height = img.height

        # Definir límites en cm
        max_width_cm = 34.65
        max_height_cm = 15.34

        # Determinar factor de escala
        if original_width > original_height:
            scale_factor = max_width_cm / (original_width * PIXEL_TO_CM)
        else:
            scale_factor = max_height_cm / (original_height * PIXEL_TO_CM)

        # Calcular nuevas dimensiones en píxeles
        new_width = int(original_width * scale_factor)
        new_height = int(original_height * scale_factor)

        # Aplicar nuevas dimensiones
        img.width = new_width
        img.height = new_height

        return img


    def _get_closest_station_name(self, station: str):
        stations = self.stations['TX_TDT'].tolist()
        found_station, score, index = process.extractOne(
            station, stations, scorer=fuzz.ratio)

        return found_station
    

    def get_station_list(self, digital_measurement_dictionary: dict):
        station_list = []
        for channel, results in digital_measurement_dictionary.items():
            if results['service_name'] in ['Caracol', 'RCN']:
                station = f"{results['station']} - CCNP".upper()
            else:
                station = f"{results['station']} - RTVC".upper()
            found_station = self._get_closest_station_name(station)
            station_list.append(found_station)

        return list(set(station_list))
    

    def plot_elevation_profile(self, lat_point: float, lon_point: float, station_list: list):

        # Coordenadas del punto de medición
        point_coord = (lat_point, lon_point)

        # Cargue del archivo DEM
        dem_file = "./resources/SRTM_30_Col1.tif"
        with rasterio.open(dem_file) as dem:
            dem_crs = dem.crs  # Obtener CRS del DEM

        # Transformador de coordenadas
        transformer = Transformer.from_crs("EPSG:4326", dem_crs.to_string(), always_xy=True)

        # Convertir coordenadas a sistema proyectado del DEM
        point_coord_projected = transformer.transform(*point_coord[::-1])

        # Número de puntos a lo largo del perfil
        number_of_points = 500

        for station in station_list:
            # Se obtienen las coordenadas de la estación desde el documento de referencias
            index = self.stations.index[self.stations['TX_TDT'] == station].tolist()[0]
            lat_station = self.stations.at[index, 'LAT_D']
            lon_station = self.stations.at[index, 'LONG_D']
            station_coord = (lat_station, lon_station)

            # Convertir coordenadas a sistema proyectado DEM
            station_coord_projected = transformer.transform(*station_coord[::-1])

            # Se crea una línea entre la estación y el punto
            line = LineString([station_coord_projected, point_coord_projected])
            interpolated_points = [line.interpolate(dist, normalized=True).coords[0] for dist in np.linspace(0, 1, number_of_points)]

            # Lista para guardar las elevaciones y distancias
            elevations = []
            distances = []

            # Abrir el DEM y procesar punto por punto
            with rasterio.open(dem_file) as dem:
                for i, (x, y) in enumerate(interpolated_points):
                    # Convertir x/y a fila/columna en el DEM
                    try:
                        row, col = dem.index(x, y)

                        # Validar que la fila y columna estén dentro de los límites del DEM
                        if 0 <= row < dem.height and 0 <= col < dem.width:
                            # Leer solo un píxel del DEM
                            elevation = dem.read(1, window=rasterio.windows.Window(col, row, 1, 1))[0, 0]
                            elevations.append(elevation)

                            # Calcular distancia acumulada
                            if i > 0:
                                prev_x, prev_y = interpolated_points[i - 1]
                                dist = np.sqrt((x - prev_x)**2 + (y - prev_y)**2) / 1000  # Convertir a km
                                distances.append(distances[-1] + dist if distances else dist)
                            else:
                                distances.append(0)
                        else:
                            print(f"Punto fuera de límites: {x}, {y}")
                            elevations.append(None)
                    except Exception as e:
                        print(f"Error procesando el punto ({x}, {y}): {e}")
                        elevations.append(None)

            # Graficar el perfil de elevación
            elevations = np.array(elevations, dtype=float)
            distances = np.array(distances, dtype=float)

            plt.figure(figsize=(13, 4))
            plt.plot(distances, np.nan_to_num(elevations), color='brown')
            plt.fill_between(distances, np.nan_to_num(elevations), color='lightcoral', alpha=0.5)
            plt.xlabel("Distancia [km]")
            plt.ylabel("Elevación [m]")
            plt.grid()
            plt.xlim(min(distances), max(distances))
            plt.ylim(min(elevations)*0.9, max(elevations)*1.05)
            plt.savefig(f"./temp/elevation_profile-{station}.png", dpi=300, bbox_inches="tight", pad_inches=0.25)  # Guarda la imagen con alta resolución


    def plot_distances_image(self, lat_point: float, lon_point: float, station_list: list):
        
        # Coordenadas del punto de medición
        point_coord = (lon_point, lat_point)

        stations = []
        for station in station_list:
            # Se obtienen las coordenadas de la estación desde el documento de referencias
            index = self.stations.index[self.stations['TX_TDT'] == station].tolist()[0]
            lat_station = self.stations.at[index, 'LAT_D']
            lon_station = self.stations.at[index, 'LONG_D']
            station_coord = (lon_station, lat_station)

            stations.append({'station_name': station, 'coordinates': station_coord})

        # Crear un GeoDataFrame para las líneas
        geometries = [LineString([point_coord, station['coordinates']]) for station in stations]

        # Creación de líneas y punto
        colors = ['blue', 'green', 'red', 'cyan', 'magenta', 'yellow', 'black', 'white']
        lines = gpd.GeoDataFrame({'color': [colors[i] for i in range(len(station_list))], 'name': station_list}, geometry=geometries, crs='EPSG:4326')
        point = gpd.GeoDataFrame({'name': ['Punto de medición']}, geometry=[Point(point_coord)], crs='EPSG:4326')

        # Cargar el archivo raster del mapa base
        raster_path = './resources/Colombia_Satelital.tif'  # Ruta al archivo raster descargado
        with rasterio.open(raster_path) as src:
            # Reproyectar las capas al CRS del raster
            lines = lines.to_crs(src.crs)
            point = point.to_crs(src.crs)

            # Recortar la región de interés automáticamente
            all_coords_proj = [point.geometry[0].coords[0]] + [lines.geometry[idx].coords[1] for idx in range(len(stations))]
            lons, lats = zip(*all_coords_proj)

            margin = max(max(lons) - min(lons), max(lats) - min(lats)) * 0.2  # Margen para los límites en las unidades del raster (metros)
            min_x, max_x = min(lons) - margin, max(lons) + margin
            min_y, max_y = min(lats) - margin, max(lats) + margin

            margin_factor = 0.2  # Factor del 20% para añadir márgenes
            x_min, x_max = min(lons), max(lons)
            y_min, y_max = min(lats), max(lats)
            margin_x = (x_max - x_min) * margin_factor
            margin_y = (y_max - y_min) * margin_factor

            # Crear la ventana (bounds -> ventana en el espacio del raster)
            window = rasterio.windows.from_bounds(min_x - margin_x, min_y - margin_y, max_x + margin_x, max_y + margin_y, src.transform)
            data = src.read(window=window)  # Leer todas las bandas de la región de interés
            transform = src.window_transform(window)  # Ajustar el transform para la ventana

            # Mostrar el mapa base recortado con colores originales
            fig, ax = plt.subplots(figsize=(10, 10))
            if data.shape[0] >= 3:  # Si el raster tiene al menos 3 bandas (RGB)
                show(data[:3], transform=transform, ax=ax)  # Mostrar las 3 primeras bandas como RGB
            else:
                show(data, transform=transform, ax=ax, cmap='gray')  # Fallback a escala de grises

            # Dibujar líneas y puntos sobre el mapa base
            lines.plot(ax=ax, color=lines['color'], linewidth=3)

            # Agregar marcadores al final de las líneas
            for idx, station in enumerate(stations):
                coord_x, coord_y = lines.geometry[idx].coords[1]  # Coordenadas finales
                ax.plot(
                    coord_x,
                    coord_y,
                    marker='*',  # Marcador circular
                    color=lines['color'][idx],
                    markersize=15,
                    label=station['station_name'],
                )

            # Dibujar el punto de medición
            coord_x_1, coord_y_1 = point.geometry[0].coords[0]

            ax.plot(
                coord_x_1,
                coord_y_1,
                marker='.',  # Marcador circular
                color='white',
                markersize=15,
                label='Punto de medición',
            )

            # Ajustar límites dinámicamente
            margin_factor = 0.2  # Factor del 10% para añadir márgenes
            x_min, x_max = min(lons), max(lons)
            y_min, y_max = min(lats), max(lats)
            margin_x = (x_max - x_min) * margin_factor
            margin_y = (y_max - y_min) * margin_factor
            ax.set_xlim(x_min - margin_x, x_max + margin_x)
            ax.set_ylim(y_min - margin_y, y_max + margin_y)
            ax.legend(loc='upper right', fontsize=10, title_fontsize=12)
            ax.axis('off')

            # Guardar el mapa como archivo de imagen
            plt.savefig('./temp/distances.png', dpi=300, bbox_inches='tight', pad_inches=0)


    def fill_register_sheet(self, site_dictionary: dict, analog_measurement_dictionary: dict):
        # Información general
        self.register_sheet['E3']  = site_dictionary['municipality']
        self.register_sheet['E4']  = site_dictionary['department']
        self.register_sheet['E5']  = site_dictionary['address']
        self.register_sheet['E6']  = site_dictionary['latitude_dms']
        self.register_sheet['E7']  = site_dictionary['longitude_dms']
        self.register_sheet['E8']  = site_dictionary['altitude']
        self.register_sheet['E9']  = site_dictionary['point']
        self.register_sheet['E10'] = site_dictionary['around']
        self.register_sheet['E11'] = site_dictionary['terrain']
        self.register_sheet['E12'] = site_dictionary['signal_path']
        self.register_sheet['E13'] = site_dictionary['signal_obstruction']

        # Servidores públicos responsables de las mediciones
        now = datetime.datetime.now() # Obetención de fecha y hora de la medida
        date = f'{str(now.day).zfill(2)}/{str(now.month).zfill(2)}/{str(now.year)}'
        self.register_sheet['L3']  = date
        self.register_sheet['H6']  = site_dictionary['engineer_1']
        self.register_sheet['H10'] = site_dictionary['engineer_2']

        # Equipo utilizado para la realización del monitoreo in situ
        self.register_sheet['R4']  = site_dictionary['instrument_type']
        self.register_sheet['T4']  = site_dictionary['instrument_brand']
        self.register_sheet['U4']  = site_dictionary['instrument_model']
        self.register_sheet['U4']  = site_dictionary['instrument_serial']

        self.register_sheet['R4']  = 'Antena'
        self.register_sheet['T4']  = site_dictionary['antenna_brand']
        self.register_sheet['U4']  = site_dictionary['antenna_model']

        # Mediciones de niveles de servicio e interferencias para televisión analógica
        for row, (channel, dic) in enumerate(analog_measurement_dictionary.items(), start=21):
            self.register_sheet[f'C{row}']  = dic['hour']
            self.register_sheet[f'D{row}']  = channel

            service_name = dic['service_name']
            if service_name == 'Caracol':
               fix_service_name = 'Caracol TV S.A.'
            elif service_name == 'RCN':
                fix_service_name = 'RCN TV S.A.'
            elif service_name == 'Canal 1':
                fix_service_name = 'Canal Uno'
            elif service_name == 'Canal Institucional':
                fix_service_name = 'C. Institucional'
            else:
                fix_service_name = service_name
            
            self.register_sheet[f'E{row}']  = fix_service_name
            self.register_sheet[f'F{row}']  = dic['station']
            self.register_sheet[f'J{row}']  = dic['power_video']
            self.register_sheet[f'K{row}']  = dic['power_audio']
            self.register_sheet[f'L{row}']  = dic['frequency_video']
            self.register_sheet[f'M{row}']  = dic['frequency_audio']


    def fill_graphical_support_sheet(self, site_dictionary: dict, analog_measurement_dictionary: dict):
        rows_for_images = [3, 22, 41, 63, 82, 101, 122, 141, 160, 181, 200, 219, 241, 260, 270, 300, 319, 338, 359, 378, 397]
        
        for index, (channel, dic) in enumerate(analog_measurement_dictionary.items()):
            municipality = site_dictionary['municipality']
            point = str(site_dictionary['point']).zfill(2)
            station = dic['station']
            service_name = dic['service_name']
            img_path = f'{municipality}/P{point}/Soportes punto de medición/{station}/CH{channel}_A_{service_name}/CH_{channel}.png'

            img = Image(img_path)
            img.width  = img.width * 0.25
            img.height = img.height  * 0.25

            self.graphical_supports_sheet.add_image(img, f'A{rows_for_images[index]}')


    def fill_general_info_sheet(self, site_dictionary: dict, digital_measurement_dictionary: dict):
        # Características del punto de medición
        self.general_info_sheet['A8']  = site_dictionary['point']
        self.general_info_sheet['C8']  = site_dictionary['municipality']
        self.general_info_sheet['M8']  = site_dictionary['department']
        self.general_info_sheet['V8']  = site_dictionary['latitude_dms']
        self.general_info_sheet['AB8'] = site_dictionary['longitude_dms']
        self.general_info_sheet['AH8'] = site_dictionary['altitude']
        self.general_info_sheet['A11'] = site_dictionary['around']
        self.general_info_sheet['H11'] = site_dictionary['terrain']
        self.general_info_sheet['M11'] = site_dictionary['signal_path']
        self.general_info_sheet['S11'] = site_dictionary['signal_obstruction']
        self.general_info_sheet['Z11'] = site_dictionary['address']

        # Equipo utilizado para la medición
        self.general_info_sheet['A15'] = site_dictionary['instrument_type']
        self.general_info_sheet['I15'] = site_dictionary['instrument_brand']
        self.general_info_sheet['Q15'] = site_dictionary['instrument_model']
        self.general_info_sheet['W15'] = site_dictionary['instrument_serial']

        self.general_info_sheet['A16'] = 'Antena'
        self.general_info_sheet['I16'] = site_dictionary['antenna_brand']
        self.general_info_sheet['Q16'] = site_dictionary['antenna_model']
																																		
        # Características de las estaciones que se reciben en el punto donde se realiza la medición
        station_list = self.get_station_list(digital_measurement_dictionary)
        for row, station in enumerate(station_list, start=20):
            self.general_info_sheet[f'A{row}'] = station
																																	
        # Perfiles de terreno desde el punto donde se realiza la medición a las estaciones monitoreadas
        rows_for_images = [72, 76, 80, 84, 88, 92, 97, 101]
        for index, station in enumerate(station_list):
            img_path = f'./temp/elevation_profile-{station}.png'

            img = Image(img_path)
            img.width  = img.width * 0.84
            img.height = img.height  * 0.84

            self.general_info_sheet.add_image(img, f'A{rows_for_images[index]}')

        # Distancias desde el punto de medición a las estaciones monitoreadas
        distance_img_path = './temp/distances.png'
        distance_img = Image(distance_img_path)
        resized_distance_img = self._resize_image(distance_img)
        self.general_info_sheet.add_image(resized_distance_img, 'A105')


    def fill_channel_sheet(self, site_dictionary: dict, digital_measurement_dictionary: dict, sfn_dictionary: dict):
        # Se crea una hoja y se llena para cada canal digital medido
        for channel, dic in digital_measurement_dictionary.items():
            self.wb_digital.copy_worksheet(self.channel_template_sheet)
            channel_sheet = self.wb_digital[self.wb_digital.sheetnames[-5]]
            channel_sheet.title = f'CH{channel}'

            # Características de la medición
            station_name = f"{dic['station']} - CCNP" if dic['service_name'] in ['Caracol', 'RCN'] else f"{dic['station']} - RTVC"
            station = self._get_closest_station_name(station_name)
            channel_sheet['A8']  = station
            channel_sheet['I8']  = dic['service_name'].upper()
            channel_sheet['Q8']  = TV_TABLE[channel]
            channel_sheet['W8']  = dic['channel_type']
            channel_sheet['AC8'] = dic['date']
            channel_sheet['AF8'] = dic['hour']

            # Características de las estaciones en red SFN que se reciben en el punto donde se realiza la medición
            try:
                sfn_stations_list = list(sfn_dictionary[channel].keys())
                sfn_stations_list.remove(dic['station'])

                for row, sfn_station in enumerate(sfn_stations_list, start=12):
                    renamed_sfn_station = f'{sfn_station} - CCNP' if dic['service_name'] in ['Caracol', 'RCN'] else f'{sfn_station} - RTVC'
                    renamed_sfn_station = self._get_closest_station_name(renamed_sfn_station)
                    channel_sheet[f'A{row}'] = renamed_sfn_station
            except KeyError:
                pass

            # Registro de verificación de cobertura y parámetros de calidad del servicio TDT. Resultados de medición por PLP.
            plps = PLP_SERVICES[dic['service_name']]
            ts_array = []
            for row, plp in enumerate(plps, start=19):
                key_plp = f'PLP_{plp}'
                channel_sheet[f'L{row}']  = dic['channel_power']
                channel_sheet[f'P{row}']  = dic[key_plp]['MRPLp']
                channel_sheet[f'R{row}']  = dic[key_plp]['BERLdpc']
                channel_sheet[f'V{row}']  = dic[key_plp]['cons']
                channel_sheet[f'Y{row}']  = dic[key_plp]['PLPCodeRate']
                channel_sheet[f'AA{row}'] = dic[key_plp]['FFTMode']
                channel_sheet[f'AD{row}'] = dic[key_plp]['GINTerval']
                channel_sheet[f'AG{row}'] = dic[key_plp]['PPATtern']
                
                try:
                    # Arreglo completo de Transport Stream
                    for i in range(len(dic[key_plp]['TS'])):
                        ts_array.append(dic[key_plp]['TS'][i])
                except KeyError:
                    pass
                
            # Análisis de Transport Stream
            if ts_array:
                for row, ts_service_result in enumerate(ts_array, start=24):
                    channel_sheet[f'M{row}']  = ts_service_result[1]
                    channel_sheet[f'O{row}']  = ts_service_result[2]
                    channel_sheet[f'Q{row}']  = ts_service_result[3]
                    channel_sheet[f'T{row}']  = ts_service_result[4]
                    channel_sheet[f'X{row}']  = ts_service_result[5]
                    channel_sheet[f'AA{row}'] = ts_service_result[6]
                    channel_sheet[f'AD{row}'] = 'No Falla'

            # Soportes de medición
            municipality = site_dictionary['municipality']
            point = str(site_dictionary['point']).zfill(2)
            station = dic['station']
            service_name = dic['service_name']
            plp = PLP_SERVICES[service_name][0]
            imgs = {
                'Channel Power': Image(f'{municipality}/P{point}/Soportes punto de medición/{station}/CH{channel}_D_{service_name}/{TV_TABLE[channel]}.png'),
                'MER':           Image(f'{municipality}/P{point}/Soportes punto de medición/{station}/CH{channel}_D_{service_name}/PLP_{plp}/{TV_TABLE[channel]}_002.png'),
                'BER':           Image(f'{municipality}/P{point}/Soportes punto de medición/{station}/CH{channel}_D_{service_name}/PLP_{plp}/{TV_TABLE[channel]}_003.png'),
                'Constelation':  Image(f'{municipality}/P{point}/Soportes punto de medición/{station}/CH{channel}_D_{service_name}/PLP_{plp}/{TV_TABLE[channel]}_001.png'),
                'Echo Pattern':  Image(f'{municipality}/P{point}/Soportes punto de medición/{station}/CH{channel}_D_{service_name}/PLP_{plp}/{TV_TABLE[channel]}_008.png'),
                'Shoulders':     Image(f'{municipality}/P{point}/Soportes punto de medición/{station}/CH{channel}_D_{service_name}/PLP_{plp}/{TV_TABLE[channel]}_010.png')
            }

            for img in imgs.values():
                img.width  = img.width/3.1
                img.height = img.height/3.1

            channel_sheet.add_image(imgs['Channel Power'], 'A44') # Se añade la imagen "ChanelPower" a la casilla A44.
            channel_sheet.add_image(imgs['MER'],           'M44') # Se añade la imagen "ModulationErrors" a la casilla M44.
            channel_sheet.add_image(imgs['BER'],           'Z44') # Se añade la imagen "DigitalOverview" a la casilla Z44.
            channel_sheet.add_image(imgs['Constelation'],  'A64') # Se añade la imagen "Constelation" a la casilla A62.
            channel_sheet.add_image(imgs['Echo Pattern'],  'M64') # Se añade la imagen "EchoPattern" a la casilla M62.
            channel_sheet.add_image(imgs['Shoulders'],     'Z64') # Se añade la imagen "Shoulders" a la casilla Z62.


    def fill_reports(self, site_dictionary: dict, analog_measurement_dictionary: dict, digital_measurement_dictionary: str, sfn_dictionary):
        # Coordenadas del punto, para las funciones de graficado
        lat_point = site_dictionary['latitude_dec']
        lon_point = site_dictionary['longitude_dec']

        # Lista de estaciones de digital
        digital_station_list = self.get_station_list(digital_measurement_dictionary)

        # Generación de gráficas de perfiles y distancias
        os.makedirs('./temp', exist_ok=True)
        self.plot_elevation_profile(lat_point, lon_point,digital_station_list)
        self.plot_distances_image(lat_point, lon_point,digital_station_list)

        # Llenado de postprocesamiento anaógico
        self.fill_register_sheet(site_dictionary, analog_measurement_dictionary)
        self.fill_graphical_support_sheet(site_dictionary, analog_measurement_dictionary)

        # Llenado del postprocesamiento tdt.
        self.fill_general_info_sheet(site_dictionary, digital_measurement_dictionary)
        self.fill_channel_sheet(site_dictionary, digital_measurement_dictionary, sfn_dictionary)

        # Guardado de archivos.
        municipality = site_dictionary['municipality']
        point = str(site_dictionary['point']).zfill(2)
        self.wb_analog.save(f'./results/{municipality}/P{point}/FOR_Registro Monitoreo In Situ TV Analógica_V0_P{point}.xlsm')
        self.wb_digital.save(f'./results/{municipality}/P{point}/FOR_Registro Monitoreo In Situ TDT_V0_P{point}.xlsm')

        shutil.rmtree('./temp')

        


if __name__ == '__main__':
    report = ExcelReport()
    # print(report.encontrar_nombre_parecido('El Tigre - RTVC'.upper()))

    diccionario_resultado = {16: {'date': '2025/03/06', 'hour': '12:08:11', 'channel_power': 68.28, 'channel_type': 'Rayleigh (σ = 3.87 dB)', 'PLP101': {'SALower': 16.7036895752, 'SAUPper': 6.03944396973, 'LEVel': 57.8406126339, 'CFOFfset': -54.9, 'BROFfset': -0.14, 'PERatio': '---', 'BERLdpc': 'ND', 'BBCH': '---', 'FERatio': '---', 'ESRatio': '---', 'GINTerval': '1/8', 'PLPCodeRate': '2/3', 'IMBalance': -3.83, 'QERRor': 0.48, 'CSUPpression': 18.2, 'MRLO': 10.384, 'MPLO': -5.534, 'MRPLp': 'ND', 'MPPLp': '---', 'ERPLp': '---', 'EPPLp': '---', 'cons': '64QAM', 'FFTMode': '16k ext', 'AMPLitude': 39.3581409454, 'PHASe': 3337.78112793, 'GDELay': 0.00020115996, 'PPATtern': 'PP3'}, 'station': 'Manjui', 'service_name': 'Caracol'},
                             17: {'date': '2025/03/06', 'hour': '12:08:11', 'channel_power': 68.28, 'channel_type': 'Rayleigh (σ = 3.87 dB)', 'PLP101': {'SALower': 16.7036895752, 'SAUPper': 6.03944396973, 'LEVel': 57.8406126339, 'CFOFfset': -54.9, 'BROFfset': -0.14, 'PERatio': '---', 'BERLdpc': 'ND', 'BBCH': '---', 'FERatio': '---', 'ESRatio': '---', 'GINTerval': '1/8', 'PLPCodeRate': '2/3', 'IMBalance': -3.83, 'QERRor': 0.48, 'CSUPpression': 18.2, 'MRLO': 10.384, 'MPLO': -5.534, 'MRPLp': 'ND', 'MPPLp': '---', 'ERPLp': '---', 'EPPLp': '---', 'cons': '64QAM', 'FFTMode': '16k ext', 'AMPLitude': 39.3581409454, 'PHASe': 3337.78112793, 'GDELay': 0.00020115996, 'PPATtern': 'PP3'}, 'station': 'Manjui', 'service_name': 'RTVC'}}

    lista_estaciones = report.get_station_list(diccionario_resultado)

    report.plot_elevation_profile(4.67760, -74.05437, lista_estaciones)
    report.plot_distances_image(4.67760, -74.05437, lista_estaciones)